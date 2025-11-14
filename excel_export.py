from __future__ import annotations
from typing import List, Dict, Optional

def export_messages_to_xlsx(path: str, results: List[Dict]) -> Optional[str]:
    """
    Eksporter søkeresultater til Excel .xlsx. Returnerer None ved suksess, ellers feiltekst.
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
        from openpyxl.styles import Alignment, Font  # type: ignore
    except Exception:
        return "openpyxl mangler. Installer med:\n  pip install openpyxl"

    wb = Workbook()
    ws = wb.active
    ws.title = "Eposter"

    headers = ["Dato", "Emne", "Fra", "Fra (smtp)", "#Vedlegg", "Ulest", "Mappe"]
    ws.append(headers)

    for r in results:
        ws.append([
            r["dt"].strftime("%Y-%m-%d %H:%M"),
            r.get("subject", ""),
            r.get("from", ""),
            r.get("from_email", ""),
            r.get("attach", 0),
            "Ja" if r.get("unread") else "",
            r.get("folder", ""),
        ])

    # topptekst
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # kolonnebredder (enkelt og effektivt)
    widths = [20, 60, 32, 32, 10, 8, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
        return None
    except Exception as e:
        return f"Feil ved lagring av Excel: {e}"
